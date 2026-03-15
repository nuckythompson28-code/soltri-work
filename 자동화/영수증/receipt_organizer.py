#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
receipt_organizer.py
1. Downloads 폴더에서 영수증 PDF 탐색
2. 바탕화면/영수증/원본/ 으로 이동 (원본 보존)
3. OCR + 텍스트 추출로 날짜/금액/업체명/품목 파싱
4. 가로모드 2장/페이지 합본 PDF 생성
5. Excel 목록 생성 → 바탕화면/영수증/ 에 저장
"""

import os
import sys
import re
import io
import shutil
from pathlib import Path
from datetime import datetime

try:
    import pdfplumber
    from pypdf import PdfReader, PdfWriter, Transformation
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import fitz          # PyMuPDF
    from PIL import Image
except ImportError as e:
    print(f"[ERROR] Missing package: {e}")
    print("Please run the .bat file to install dependencies.")
    input("Press any key to exit...")
    sys.exit(1)

# ── OCR 백엔드 선택 (Windows 내장 우선 → easyocr 폴백)
OCR_ENGINE = None

def _try_load_easyocr():
    try:
        import easyocr
        return "easyocr"
    except Exception:
        return None

# PowerShell OCR은 Windows 10/11에 내장 — 항상 시도
OCR_ENGINE = _try_load_easyocr() or "powershell"
print(f"[OCR] Engine: {OCR_ENGINE}")

# ─────────────────────────────────────────────
# 경로 설정
# ─────────────────────────────────────────────
DOWNLOADS     = Path.home() / "Downloads"
SCRIPT_DIR    = Path(__file__).parent
ORIGINALS_DIR = SCRIPT_DIR / "원본"
OUTPUT_PDF    = SCRIPT_DIR / "영수증_합본.pdf"
OUTPUT_XLSX   = SCRIPT_DIR / "영수증_목록.xlsx"

RECEIPT_PATTERNS = [
    "Npay_카드영수증_*.pdf",
    "Npay_카드영수증일괄_*.pdf",
    "Npay_현금영수증_*.pdf",
    "receipt_*.pdf",
    "Receipt*.pdf",
    "신용카드영수증_*.pdf",
    "영수증_*.pdf",
    "현금영수증*.pdf",
]
EXCLUDE_KEYWORDS = [
    "원천징수", "소득자보관", "발행자보고", "사용내역", "근로소득"
]

A4_W, A4_H = landscape(A4)

# ─────────────────────────────────────────────
# 1. 파일 수집 & 이동
# ─────────────────────────────────────────────
def collect_and_move(downloads: Path, dest: Path) -> list:
    dest.mkdir(parents=True, exist_ok=True)
    found = set()
    for pattern in RECEIPT_PATTERNS:
        for f in downloads.glob(pattern):
            if not any(kw in f.name for kw in EXCLUDE_KEYWORDS):
                found.add(f)

    moved = []
    for src in sorted(found):
        dst = dest / src.name
        if not dst.exists():
            shutil.move(str(src), str(dst))
            print(f"  Moved: {src.name}")
        else:
            print(f"  Already in dest: {src.name}")
        moved.append(dst)

    # 이미 원본 폴더에 있는 파일 포함
    for pattern in RECEIPT_PATTERNS:
        for f in dest.glob(pattern):
            if not any(kw in f.name for kw in EXCLUDE_KEYWORDS):
                if f not in moved:
                    moved.append(f)

    def sort_key(p):
        d = date_from_filename(p.name)
        return d if d else "0000-00-00"

    return sorted(set(moved), key=sort_key, reverse=True)


# ─────────────────────────────────────────────
# 2. PDF → 이미지 추출
# ─────────────────────────────────────────────
def pdf_to_pil_images(pdf_path: Path) -> list:
    """PDF 각 페이지를 PIL Image 리스트로 반환 (2× 확대)"""
    images = []
    try:
        doc = fitz.open(str(pdf_path))
        for page in doc:
            # 방법 1: 임베드된 이미지 직접 추출
            img_list = page.get_images(full=True)
            if img_list:
                xref = img_list[0][0]
                base_img = doc.extract_image(xref)
                pil = Image.open(io.BytesIO(base_img["image"])).convert("RGB")
                images.append(pil)
            else:
                # 방법 2: 페이지를 렌더링
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                images.append(pil)
        doc.close()
    except Exception as e:
        print(f"    [img] Error: {e}")
    return images


# ─────────────────────────────────────────────
# 3. OCR
# ─────────────────────────────────────────────
def _ocr_powershell(pil_image: Image.Image) -> str:
    """Windows 10/11 내장 OCR — 임시파일 경유 PowerShell"""
    import subprocess, tempfile, os

    tmp_img = tmp_ps = None
    try:
        # 이미지를 임시 파일로 저장 (명령줄 길이 제한 우회)
        fd, tmp_img = tempfile.mkstemp(suffix=".png")
        os.close(fd)
        pil_image.save(tmp_img, format="PNG")

        ps_script = """
param([string]$ImgPath)
try {
    Add-Type -AssemblyName System.Runtime.WindowsRuntime
    $null = [Windows.Storage.StorageFile, Windows.Storage, ContentType=WindowsRuntime]
    $null = [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]
    $null = [Windows.Graphics.Imaging.BitmapDecoder, Windows.Foundation, ContentType=WindowsRuntime]

    function Await($Task) {
        $t = [System.WindowsRuntimeSystemExtensions]::AsTask($Task)
        $t.Wait(-1) | Out-Null; return $t.Result
    }

    $file   = Await([Windows.Storage.StorageFile]::GetFileFromPathAsync($ImgPath))
    $stream = Await($file.OpenReadAsync())
    $dec    = Await([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream))
    $bmp    = Await($dec.GetSoftwareBitmapAsync())

    $lang   = [Windows.Globalization.Language]::new("ko")
    $engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromLanguage($lang)
    if (-not $engine) { $engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages() }
    if (-not $engine) { exit 0 }

    $res = Await($engine.RecognizeAsync($bmp))
    Write-Output $res.Text
} catch {
    Write-Error $_.Exception.Message
    exit 1
}
"""
        # PS 스크립트도 파일로 저장
        fd2, tmp_ps = tempfile.mkstemp(suffix=".ps1")
        os.close(fd2)
        with open(tmp_ps, "w", encoding="utf-8") as f:
            f.write(ps_script)

        result = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive",
             "-ExecutionPolicy", "Bypass",
             "-File", tmp_ps, tmp_img],
            capture_output=True, text=True, timeout=60
        )
        text = result.stdout.strip()
        if result.returncode != 0 and result.stderr:
            print(f"    [ps-ocr] {result.stderr.strip()[:120]}")
        return text

    except Exception as e:
        print(f"    [ps-ocr] {e}")
        return ""
    finally:
        for p in (tmp_img, tmp_ps):
            if p and os.path.exists(p):
                try: os.unlink(p)
                except: pass


def ocr_with_easyocr(pil_image: Image.Image) -> str:
    import easyocr
    import numpy as np
    try:
        reader = easyocr.Reader(['ko', 'en'], verbose=False)
        result = reader.readtext(np.array(pil_image), detail=0, paragraph=True)
        return "\n".join(result)
    except Exception as e:
        print(f"    [easyocr] {e}")
        return ""


def run_ocr(pdf_path: Path) -> str:
    images = pdf_to_pil_images(pdf_path)
    if not images:
        return ""
    all_text = []
    for img in images:
        if OCR_ENGINE == "easyocr":
            t = ocr_with_easyocr(img)
        else:
            t = _ocr_powershell(img)   # powershell (default)
        if t:
            all_text.append(t)
    return "\n".join(all_text)


# ─────────────────────────────────────────────
# 4. 텍스트 추출
# ─────────────────────────────────────────────
def extract_text_pdf(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            parts = [p.extract_text() for p in pdf.pages if p.extract_text()]
            return "\n".join(parts).strip()
    except Exception:
        return ""


# ─────────────────────────────────────────────
# 5. 날짜 – 파일명 폴백
# ─────────────────────────────────────────────
def date_from_filename(filename: str) -> str:
    m = re.search(r'[_\-](\d{4})(\d{2})(\d{2})', filename)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        if 2000 <= int(y) <= 2035 and 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            return f"{y}-{mo}-{d}"
    return ""


# ─────────────────────────────────────────────
# 6. 정보 파싱
# ─────────────────────────────────────────────
def parse_info(text: str, filename: str) -> dict:
    info = {"date": "", "amount": "", "merchant": "", "items": ""}

    # 날짜
    for pat in [
        r'거래일자\s*[:\s]*(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'승인일시\s*[:\s]*(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'이용일자\s*[:\s]*(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'결제일시\s*[:\s]*(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'날\s*짜\s*[:\s]*(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'(\d{4}년\s*\d{1,2}월\s*\d{1,2}일)',
        r'(\d{4}[/\-.]\d{2}[/\-.]\d{2})',
        r'(\d{4}\.\d{2}\.\d{2})',
    ]:
        m = re.search(pat, text)
        if m:
            info["date"] = m.group(1).strip()
            break
    if not info["date"]:
        info["date"] = date_from_filename(filename)

    # 금액
    for pat in [
        r'결제금액\s*[:\s]*[￦₩]?\s*([\d,]+)',
        r'합계금액\s*[:\s]*[￦₩]?\s*([\d,]+)',
        r'총\s*금액\s*[:\s]*[￦₩]?\s*([\d,]+)',
        r'승인금액\s*[:\s]*[￦₩]?\s*([\d,]+)',
        r'이용금액\s*[:\s]*[￦₩]?\s*([\d,]+)',
        r'결제\s*금액[^\d]*?([\d,]+)\s*원',
        r'[￦₩]\s*([\d,]+)',
        r'([\d,]+)\s*원\b',
    ]:
        m = re.search(pat, text)
        if m:
            val = m.group(1).replace(",", "")
            if val.isdigit() and 100 <= int(val) <= 99999999:
                info["amount"] = val
                break

    # 업체명
    for pat in [
        r'가맹점명\s*[:\s]*([^\n\r]+?)(?:\s{2,}|카드잔액|대표자|$)',
        r'이용가맹점\s*[:\s]*([^\n\r]+?)(?:\s{2,}|$)',
        r'상\s*호\s*[:\s：]*([^\n\r]+)',
        r'업체명\s*[:\s：]*([^\n\r]+)',
        r'사용처\s*[:\s：]*([^\n\r]+)',
        r'판매자\s*[:\s：]*([^\n\r]+)',
        r'가맹점\s+([가-힣A-Za-z0-9\(\)&\-\.\s]+?)(?:\n|$)',
    ]:
        m = re.search(pat, text, re.MULTILINE)
        if m:
            val = re.split(r'\s{3,}|\t', m.group(1).strip())[0].strip()
            if val and 2 <= len(val) <= 60:
                info["merchant"] = val
                break

    # 품목/상품명
    item_lines = []
    for pat in [
        r'품\s*명\s*[:\s：]*([^\n\r]+)',
        r'상품명\s*[:\s：]*([^\n\r]+)',
        r'제품명\s*[:\s：]*([^\n\r]+)',
        r'항\s*목\s*[:\s：]*([^\n\r]+)',
        r'이용내역\s*[:\s：]*([^\n\r]+)',
        r'구매품목\s*[:\s：]*([^\n\r]+)',
        r'메뉴\s*[:\s：]*([^\n\r]+)',
    ]:
        for m in re.finditer(pat, text, re.MULTILINE):
            val = m.group(1).strip()
            if val and 2 < len(val) < 80:
                item_lines.append(val)
    if item_lines:
        info["items"] = " / ".join(list(dict.fromkeys(item_lines))[:3])

    return info


# ─────────────────────────────────────────────
# 7. PDF 합본
# ─────────────────────────────────────────────
def create_blank_landscape():
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=landscape(A4))
    c.setFillColorRGB(1, 1, 1)
    c.rect(0, 0, A4_W, A4_H, fill=1, stroke=0)
    c.setStrokeColorRGB(0.75, 0.75, 0.75)
    c.setLineWidth(0.5)
    c.line(A4_W / 2, 12, A4_W / 2, A4_H - 12)
    c.save(); packet.seek(0)
    return PdfReader(packet).pages[0]


def place_on_slot(output_page, source_page, slot: int, margin: float = 12):
    half_w = A4_W / 2
    avail_w = half_w - margin * 2
    avail_h = A4_H - margin * 2
    src_w = float(source_page.mediabox.width)
    src_h = float(source_page.mediabox.height)
    scale = min(avail_w / src_w, avail_h / src_h)
    x = slot * half_w + margin + (avail_w - src_w * scale) / 2
    y = margin + (avail_h - src_h * scale) / 2
    output_page.merge_transformed_page(
        source_page, Transformation().scale(scale).translate(x, y), expand=False)


def create_combined_pdf(receipt_files: list, output_path: Path):
    print("\n[PDF] Creating combined PDF...")
    all_pages = []
    for rf in receipt_files:
        try:
            reader = PdfReader(str(rf))
            for page in reader.pages:
                all_pages.append((rf.name, page))
        except Exception as e:
            print(f"  [WARN] {rf.name}: {e}")

    writer = PdfWriter()
    for i in range(0, len(all_pages), 2):
        blank = create_blank_landscape()
        place_on_slot(blank, all_pages[i][1], 0)
        if i + 1 < len(all_pages):
            place_on_slot(blank, all_pages[i + 1][1], 1)
        writer.add_page(blank)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    save_path = output_path
    try:
        with open(save_path, "wb") as f:
            writer.write(f)
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        save_path = output_path.parent / f"영수증_합본_{ts}.pdf"
        with open(save_path, "wb") as f:
            writer.write(f)
        print(f"\n  [!] 기존 PDF가 열려 있어 새 파일로 저장됨: {save_path.name}")
        print("      기존 PDF 창을 닫고 다음 번엔 동일 파일에 덮어씁니다.")
    print(f"  {len(all_pages)} receipts -> {len(writer.pages)} pages -> {save_path.name}")


# ─────────────────────────────────────────────
# 8. Excel
# ─────────────────────────────────────────────
def create_excel(receipt_files: list, output_path: Path):
    print("\n[Excel] Building spreadsheet...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영수증목록"

    thin = Side(style="thin", color="C0C8D8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1,
        value=f"영수증 목록  ─  총 {len(receipt_files)}건  │  {datetime.now().strftime('%Y년 %m월 %d일')} 생성")
    t.font = Font(name="Arial", bold=True, size=13, color="1B3A6B")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.fill = PatternFill("solid", start_color="D3E4FF", end_color="D3E4FF")
    ws.row_dimensions[1].height = 30

    headers = ["No.", "파일명", "날짜", "금액 (원)", "업체명/가맹점", "품목/내역", "비고"]
    col_widths = [5, 44, 14, 14, 30, 36, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = PatternFill("solid", start_color="2A5298", end_color="2A5298")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 20

    fills = [
        PatternFill("solid", start_color="F4F8FF", end_color="F4F8FF"),
        PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF"),
    ]

    total_amount = 0
    ocr_count = 0
    manual_count = 0

    for idx, rf in enumerate(receipt_files, 1):
        text = extract_text_pdf(rf)
        source = "text"

        if not text:
            if OCR_ENGINE:
                print(f"  [{idx:2}/{len(receipt_files)}] OCR: {rf.name}")
                text = run_ocr(rf)
                if text:
                    source = "ocr"
                    ocr_count += 1
            if not text:
                print(f"  [{idx:2}/{len(receipt_files)}] manual: {rf.name}")
                manual_count += 1
        else:
            print(f"  [{idx:2}/{len(receipt_files)}] text: {rf.name}")

        info = parse_info(text, rf.name)

        if source == "ocr":
            note = "OCR"
        elif not text:
            note = "수동입력필요"
        else:
            note = ""

        row = idx + 2
        fill = fills[idx % 2]
        amount_val = int(info["amount"]) if info["amount"].isdigit() else None
        if amount_val:
            total_amount += amount_val

        row_data = [idx, rf.name, info["date"] or "-", amount_val,
                    info["merchant"] or "-", info["items"] or "-", note]
        aligns = ["center", "left", "center", "right", "left", "left", "center"]

        for col, (val, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws.cell(row=row, column=col, value=val)
            is_manual = (col == 7 and val == "수동입력필요")
            cell.font = Font(name="Arial", size=10,
                             color="C0392B" if is_manual else "000000")
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                       wrap_text=(col == 6))
            if col == 4 and val is not None:
                cell.number_format = "#,##0"
        ws.row_dimensions[row].height = 18

    tot_row = len(receipt_files) + 3
    for col in range(1, 8):
        cell = ws.cell(row=tot_row, column=col)
        cell.fill = PatternFill("solid", start_color="E8F0FE", end_color="E8F0FE")
        cell.font = Font(name="Arial", bold=True, size=11, color="1B3A6B")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=tot_row, column=1, value="합계")
    if total_amount:
        c = ws.cell(row=tot_row, column=4, value=total_amount)
        c.number_format = "#,##0"
    parts = []
    if ocr_count:
        parts.append(f"OCR: {ocr_count}건")
    if manual_count:
        parts.append(f"수동입력: {manual_count}건")
    ws.cell(row=tot_row, column=7, value="  ".join(parts))
    ws.row_dimensions[tot_row].height = 22
    ws.freeze_panes = "A3"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 파일이 열려있으면 타임스탬프 이름으로 저장
    save_path = output_path
    try:
        wb.save(str(save_path))
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        save_path = output_path.parent / f"영수증_목록_{ts}.xlsx"
        wb.save(str(save_path))
        print(f"\n  [!] 기존 파일이 열려 있어 새 파일로 저장됨: {save_path.name}")
        print("      기존 Excel 창을 닫고 다음 번엔 동일 파일에 덮어씁니다.")

    print(f"\n  Saved: {save_path}")
    if total_amount:
        print(f"  Total: {total_amount:,} KRW")
    print(f"  OCR: {ocr_count} | Manual: {manual_count}")


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  Receipt Organizer")
    print("=" * 55)

    if not DOWNLOADS.exists():
        print(f"\n[ERROR] Downloads not found: {DOWNLOADS}")
        input("Press any key to exit...")
        sys.exit(1)

    print(f"\nSource:  {DOWNLOADS}")
    print(f"Archive: {ORIGINALS_DIR}")
    print()

    receipts = collect_and_move(DOWNLOADS, ORIGINALS_DIR)
    if not receipts:
        print("[ERROR] No receipt files found.")
        input("Press any key to exit...")
        sys.exit(1)

    print(f"\nTotal: {len(receipts)} receipts")
    create_combined_pdf(receipts, OUTPUT_PDF)
    create_excel(receipts, OUTPUT_XLSX)

    print("\n" + "=" * 55)
    print("  DONE!")
    print(f"  Originals -> {ORIGINALS_DIR}")
    print(f"  PDF       -> {OUTPUT_PDF}")
    print(f"  Excel     -> {OUTPUT_XLSX}")
    print("=" * 55)
    input("\nPress any key to close...")


if __name__ == "__main__":
    main()
